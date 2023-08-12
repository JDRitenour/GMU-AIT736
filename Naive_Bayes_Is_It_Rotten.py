# references
# https://scikit-learn.org/stable/modules/naive_bayes.html
# https://www.tutorialspoint.com/how-to-build-naive-bayes-classifiers-using-python-scikit-learn

# libraries imports
# from datetime import date
import time
import datetime
import os
import numpy as np
import openpyxl
import tracemalloc
from datetime import datetime
from PIL import Image
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import GaussianNB
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix
from openpyxl import Workbook

# varaiables
start_time = datetime.now()
train_start_time = None
test_start_time = None

complete_time = None
train_complete_time = None
test_complete_time = None

total_run_time = None
total_train_time = None
total_test_time = None

model_name = 'Naive Bayes'
log_file = model_name+"ML_.csv"
# dataset_train = r"D:\AIT 736\Group Project\archive\dataset\Train"
# dataset_test = r"D:\AIT 736\Group Project\archive\dataset\Test"
dataset = r"D:\AIT 736\Group Project\archive\dataset2"
dataset_test_pct = 0.2
dataset_random_state = 42
wb = openpyxl.load_workbook(r"D:\AIT 736\Group Project\Is_It_Rotten\IsItRotten.xlsx")
ws = wb[r"Run Result Data"]
ws['B3'] = model_name

# start logging
#f = open(log_file, "a")
#f.write('Start time ≈ '+start_time.strftime("%m/%d/%Y, %H:%M:%S"))
#f.close()
#Start logging
print("Running classifier: "+model_name)
print("Started at: ", start_time.year, start_time.month, start_time.day, start_time.hour, start_time.minute, start_time.second, start_time.microsecond)
print("using data set: "+dataset)
if dataset == r"D:\AIT 736\Group Project\archive\dataset2":
    print("data set testing reserve is: ",dataset_test_pct)

# Images must be converted into feature vectors
# Create a color histogram which represents the distribution of the composition of colors in the image
# Function creates histogram for image
def get_color_histogram(image_path, rgb_bins=(8, 8, 8)):
    # initialize histogram b
    histogram = []
    # open image
    image = Image.open(image_path)
    #print('opening: '+image_path)
    # standardize image size
    image = image.resize((512, 512))
    #print('resizing: ' + image_path)
    # image to an array
    image_array = np.array(image)
    for color_channel in range(3):  # looping through the color channels (R, G, B)
        channel_hist, _ = np.histogram(image_array[:, :, color_channel], bins=rgb_bins[color_channel], range=(0, 512))
        histogram.extend(channel_hist)
    #print('histogram: ',histogram)
    return histogram

# Function creates lists of histograms and labels for images
def create_image_histograms_and_labels(dir):
    # list of image histograms
    image_histograms = []
    # list of labels
    image_labels = []
    # iterate through folder of images and set the folder name as the label
    for subdir in os.listdir(dir):
        path = os.path.join(dir, subdir)
        for image in os.listdir(path):
            image_path = os.path.join(path, image)
            # get the histogram for the image
            hist = get_color_histogram(image_path)
            #print('histogram: ',hist)
            image_histograms.append(hist)
            # extract label from path
            image_dir = os.path.basename(path)
            #print('label path: '+image_dir)
            if image_dir[:1] == 'f':
                label = 'fresh'
                #print(label)
            elif image_dir[:1] == 'r':
                label = 'rotten'
                #print(label)
            # store in lists
            image_labels.append(label)
    #print(np.array(image_histograms), np.array(image_labels))
    return np.array(image_histograms), np.array(image_labels)

# For singular data set
# Create lists of histograms and labels for images
print('Create lists of histograms and labels')
X, y = create_image_histograms_and_labels(dataset)
print('Histograms and labels ready')

# Split the dataset into train and test sets with dataset_test_pct and dataset_random_state
print('Splitting and randomizing the dataset')
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=dataset_test_pct, random_state=dataset_random_state)
print('Splitting and randomizing complete')

# For separate train and test datasets
# Create lists of histograms and labels for images
# X_train, y_train = create_image_histograms_and_labels(dataset_train)
# X_test, y_test = create_image_histograms_and_labels(dataset_test)

# Split the dataset into train and test sets with dataset_test_pct and dataset_random_state
# X_train, y_train = np.array(X_train), np.array(y_train)
# X_test, y_test = np.array(X_test), np.array(y_test)

# Naive Bayes classifier training
print('training with '+model_name)
tracemalloc.start()
train_start_time = datetime.now()
Naive_Bayes_classifier = GaussianNB()
Naive_Bayes_classifier.fit(X_train, y_train)
train_complete_time = datetime.now()
train_memory = tracemalloc.get_traced_memory()
tracemalloc.stop()
total_train_time = train_complete_time - train_start_time
ws['C3'] = train_start_time
ws['D3'] = train_complete_time
ws['E3'] = total_train_time
print('training complete')
print('train time', total_train_time)
ws['I3'] = train_memory[0]
ws['J3'] = train_memory[1]
# print('train memory', train_memory)
# print('train memory', train_memory[0])
# print('train memory', train_memory[1])

# Naive Bayes classifier testing
print('testing '+model_name)
tracemalloc.start()
test_start_time = datetime.now()
y_predict = Naive_Bayes_classifier.predict(X_test)
test_complete_time = datetime.now()
test_memory = tracemalloc.get_traced_memory()
tracemalloc.stop()
total_test_time = test_complete_time - test_start_time
print('testing complete')
ws['F3'] = train_start_time
ws['G3'] = train_complete_time
ws['H3'] = total_train_time
ws['L3'] = test_memory[0]
ws['M3'] = test_memory[1]

# ML Model Eval
accuracy = accuracy_score(y_test, y_predict)
conf_matrix = confusion_matrix(y_test, y_predict)
classification_rep = classification_report(y_test, y_predict)
ws['O3'] = accuracy

#print(f"Accuracy: {accuracy:.2f}")
#print("Confusion Matrix:")
#print(conf_matrix)
#print("Classification Report:")
#print(classification_rep)

#  calculate runtime
complete_time = datetime.now()
total_run_time = complete_time - start_time
print(model_name+' completed at: ', total_run_time)

wb.save(r"D:\AIT 736\Group Project\Is_It_Rotten\IsItRotten.xlsx")
quit()