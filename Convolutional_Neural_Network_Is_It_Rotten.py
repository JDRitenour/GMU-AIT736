# references
# https://scikit-learn.org/stable/modules/neural_networks_supervised.html
# https://www.tensorflow.org/tutorials/images/cnn
# https://www.datacamp.com/tutorial/convolutional-neural-networks-python

# libraries imports
# from datetime import date
import time
import datetime
import os
import numpy as np
import openpyxl
import tensorflow as tf
from datetime import datetime
from PIL import Image
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
from sklearn.metrics import accuracy_score, classification_report, confusion_matrix
from openpyxl import Workbook

# varaiables
start_time = datetime.now()
train_start_time = None
test_start_time = None

complete_time = None
train_complete_time = None
test_complete_time = None

total_run_time = None
total_train_time = None
total_test_time = None

model_name = 'Convolutional_Neural_Network'
log_file = model_name+"ML_.csv"
# dataset_train = r"D:\AIT 736\Group Project\archive\dataset\Train"
# dataset_test = r"D:\AIT 736\Group Project\archive\dataset\Test"
dataset = r"D:\AIT 736\Group Project\archive\dataset2"
dataset_test_pct = 0.2
dataset_random_state = 42
wb = openpyxl.load_workbook(r"D:\AIT 736\Group Project\Is_It_Rotten\IsItRotten.xlsx")
ws = wb[r"Run Result Data"]
ws['B5'] = model_name

# start logging
#f = open(log_file, "a")
#f.write(start_time)
#f.close()
#Start logging
print("Running classifier: "+model_name)
print("Started at: ", start_time.year, start_time.month, start_time.day, start_time.hour, start_time.minute, start_time.second, start_time.microsecond)
print("using data set: "+dataset)
if dataset == r"D:\AIT 736\Group Project\archive\dataset2":
    print("data set testing reserve is: ",dataset_test_pct)

# Does not need histogram as it uses pixel value

def create_image_and_label_lists(dir):
    # list of images
    images = []
    # list of labels
    image_labels = []
    # iterate through folder of images and set the folder name as the label
    for subdir in os.listdir(dir):
        path = os.path.join(dir, subdir)
        for image in os.listdir(path):
            image_path = os.path.join(path, image)
            # standardize image size
            image = Image.open(image_path).resize((128, 128))
            if image.mode != 'RGB':
                image = image.convert('RGB')

            # image = image.resize((128, 128))
            # image to an array with pixel scaling
            image_array = np.array(image) / 255
            # store in image list
            images.append(np.array(image))
            # extract label from path
            image_dir = os.path.basename(path)
            # print('label path: '+image_dir)
            if image_dir[:1] == 'f':
                label = 'fresh'
                # print(label)
            elif image_dir[:1] == 'r':
                label = 'rotten'
                # print(label)
            # store in lists
            image_labels.append(label)
    #print(np.array(images)), np.array(image_labels))
    return np.array(images), np.array(image_labels)

# For singular data set
# Create lists of images and labels
print('Create lists of Images and labels')
X, y = create_image_and_label_lists(dataset)
print('Images and labels ready')

# Assign labels numerical values
print('Assign labels numerical values')
label_encoder = LabelEncoder()
y_encoded = label_encoder.fit_transform(y)
print('Assignment complete')

# Split the dataset into train and test sets with dataset_test_pct and dataset_random_state
print('Splitting and randomizing the dataset')
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=dataset_test_pct, random_state=dataset_random_state)
print('Splitting and randomizing complete')
print('y_train:')
print(y_train)
print('y_test:')
print(y_test)

# Convolutional Neural Network model
# tf.keras (user friendly), but there are alternatives Plain TensorFlow API, Estimators API, tf.Module, tf.function, TF-Agents, TFLite, TFX
# Rectified Linear Unit(RELU) or alternatives Leaky ReLU, Parametric ReLU, Exponential Linear Unit, Scaled Exponential Linear Unit
# 2D convolutional layer (set of filters whose parameters need to be learned) with size of 3x3 pixels & input images 128, 128, 3
# Max-pooling is (2, 2)
# Convolutional Neural Network architecture involves stacking multiple convolutional layers followed by pooling layers to progressively reduce the spatial dimensions
model = tf.keras.Sequential([
    tf.keras.layers.Conv2D(32, (3, 3), activation='relu', input_shape=(128, 128, 3)),
    tf.keras.layers.MaxPooling2D((2, 2)),
    tf.keras.layers.Conv2D(64, (3, 3), activation='relu'),
    tf.keras.layers.MaxPooling2D((2, 2)),
    tf.keras.layers.Conv2D(128, (3, 3), activation='relu'),
    tf.keras.layers.MaxPooling2D((2, 2)),
    tf.keras.layers.Flatten(),
    tf.keras.layers.Dense(128, activation='relu'),
    tf.keras.layers.Dense(len(label_encoder.classes_), activation='softmax')
])

# Compile
# optimizer Adaptive Moment Estimation (adam)
# Alternatives Stochastic Gradient Descent (SGD), SGD with Momentum, Adagrad (Adaptive Gradient Algorithm), RMSprop (Root Mean Square Propagation), Adadelta, Nesterov Accelerated Gradient (NAG),AdamW,L-BFGS (Limited-memory Broyden-Fletcher-Goldfarb-Shanno),Ranger
# sparse_categorical_crossentropy Computes the sparse categorical crossentropy loss
# evaluation metrics to be used during training and evaluation
model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['accuracy'])

# Convolutional Neural Network training
print('training with '+model_name)
train_start_time = datetime.now()
model.fit(X_train, y_train, epochs=10, batch_size=32, validation_split=0.2)
train_complete_time = datetime.now()
total_train_time = train_complete_time - train_start_time
ws['C5'] = total_train_time
print('training complete')
print('train time', total_train_time)

# Convolutional Neural Network testing
print('testing '+model_name)
test_start_time = datetime.now()
y_predict_encoded = model.predict(X_test).argmax(axis=1)
y_predict = label_encoder.inverse_transform(y_predict_encoded)
test_complete_time = datetime.now()
total_test_time = test_complete_time - test_start_time
print('testing complete')

# ML Model Eval
accuracy = accuracy_score(y_test, y_predict)
conf_matrix = confusion_matrix(y_test, y_predict)
classification_rep = classification_report(y_test, y_predict)
ws['D5'] = accuracy

#print(f"Accuracy: {accuracy:.2f}")
#print("Confusion Matrix:")
#print(conf_matrix)
#print("Classification Report:")
#print(classification_rep)

wb.save(r"D:\AIT 736\Group Project\Is_It_Rotten\IsItRotten.xlsx")